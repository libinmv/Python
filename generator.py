
from dataclasses import field
import pandas as pd
from os import path
import openpyxl as xl
import formulas
import pdfkit

spreadsheet = "/Users/libin/Downloads/sample_2.xlsx"
# Provided column names remain consistent, Total Cost to Company is taken as input

wb_init = xl.load_workbook(filename=spreadsheet, data_only=False)
ws = wb_init.active

num_Fixed = int(input("Enter number of fixed fields "))

for i in range(num_Fixed):
    fixed_field = str(input("Enter Fixed Field name - " + str(i+1) + ": "))
    fixed_value = int(input("Enter value of Fixed Field - " + str(i+1) + ": "))
    columnL = str(input("Enter B if monthly fixed value; else enter C: "))
    row_num = 0
    for cell in ws['A']:
        row_num += 1
        if cell.value == fixed_field:
            print(ws['B' + str(row_num)].value)
            ws[columnL + str(row_num)] = fixed_value
            print(ws[columnL + str(row_num)].value)
            print(ws['B' + str(row_num)].value)
            break
# input_CTC = 250000
# field_name = "Base Salary"
wb_init.save(spreadsheet)
wb_init.close
fpath = spreadsheet
dirname = path.dirname(fpath)
xl_model = formulas.ExcelModel().loads(fpath).finish()
xl_model.calculate()
xl_model.write(dirpath=dirname)
wb = xl.load_workbook(filename=spreadsheet, data_only=True)
ws = wb.active
# print(ws['B20'].value)
wb.save(spreadsheet)
df = pd.read_excel(spreadsheet)
# df.to_html("file.html")
# pdfkit.from_file("file.html", "file.pdf")
print(df)
